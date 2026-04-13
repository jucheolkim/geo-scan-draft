"""
GEO Scan v3 — CREAMWORKS
브랜드 AI 검색 진단 시스템

흐름:
  STEP 1. 브랜드 정보 입력
  STEP 2. 웹 리서치 + 질문 7개 자동 생성 (Claude API)
  STEP 3. 질문 확인 / 수정
  STEP 4. AI 답변 수집 (GPT OFF / GPT ON / Gemini)
  STEP 5. Excel 저장 → 클로드 프로젝트로 전달

Excel 3시트:
  - 브랜드 정보
  - AI 답변 수집  (답변 원문 전체 + 언급 여부 + 언급 맥락)
  - B2A 분석 결과
"""

import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl.utils
import io
import json
import re
from datetime import datetime

# ── 색상 상수 ─────────────────────────────────
CW_PURPLE       = "#7030A0"
CW_PURPLE_LIGHT = "#EADCF4"
CW_PURPLE_MID   = "#7C5CBF"

# ── 페이지 설정 ───────────────────────────────
st.set_page_config(
    page_title="GEO Scan — CREAMWORKS",
    page_icon="💜",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown(f"""
<style>
  .main .block-container {{ padding-top: 1.2rem; max-width: 900px; margin: 0 auto; }}

  .app-header {{
      background: linear-gradient(135deg, #0f0f0f 55%, #7030A055);
      padding: 22px 28px; border-radius: 14px;
      margin-bottom: 1.4rem; color: white;
  }}
  .app-header .label  {{ font-size:.75rem; color:#888; letter-spacing:2px; font-weight:600; }}
  .app-header .title  {{ font-size:1.65rem; font-weight:800; margin:6px 0 4px; }}
  .app-header .subtitle {{ font-size:.84rem; color:#bbb; }}

  .step-label {{ font-size:.72rem; font-weight:600; color:#999; letter-spacing:1px; text-transform:uppercase; margin-bottom:2px; }}
  .step-title {{ font-size:1.1rem; font-weight:700; color:#1a1a1a; margin-bottom:.8rem; }}

  .cw-box {{
      background:{CW_PURPLE_LIGHT}; border-left:4px solid {CW_PURPLE};
      padding:13px 17px; border-radius:0 10px 10px 0;
      margin:8px 0 14px; font-size:.87rem; color:#3a2d7a;
  }}
  .info-box {{
      background:#f0f4ff; border:1px solid #c7d3ff; border-radius:10px;
      padding:12px 17px; font-size:.87rem; color:#2c3e80; margin-bottom:.9rem;
  }}
  .warn-box {{
      background:#fff8e1; border:1px solid #ffe082; border-radius:10px;
      padding:12px 17px; font-size:.87rem; color:#6d4c00; margin-bottom:.9rem;
  }}

  .q-card {{
      border:1.5px solid #e8e0f0; border-radius:10px;
      padding:15px 17px; margin-bottom:10px; background:white;
  }}
  .q-badge {{
      display:inline-block; border-radius:6px;
      padding:3px 10px; font-size:.78rem; font-weight:700;
      color:white; margin-right:8px;
  }}
  .stage-badge {{
      display:inline-block; background:#f0ecf8; color:{CW_PURPLE};
      border-radius:6px; padding:2px 8px; font-size:.72rem; font-weight:600;
  }}

  .mention-yes {{ background:#d4edda; color:#155724; padding:2px 9px; border-radius:8px; font-size:.76rem; font-weight:700; }}
  .mention-no  {{ background:#f8d7da; color:#721c24; padding:2px 9px; border-radius:8px; font-size:.76rem; font-weight:700; }}
  .mention-empty {{ background:#f0f0f0; color:#888; padding:2px 9px; border-radius:8px; font-size:.76rem; }}

  .divider {{ border:none; border-top:1px solid #eee; margin:1.4rem 0; }}

  /* 스텝 진행바 */
  .step-bar {{ display:flex; gap:0; margin-bottom:1.4rem; }}
  .step-item {{ flex:1; text-align:center; padding:8px 4px; font-size:.72rem; font-weight:600; }}
  .step-active {{ background:{CW_PURPLE}; color:white; }}
  .step-done   {{ background:#27ae60; color:white; }}
  .step-todo   {{ background:#f0f0f0; color:#aaa; }}
  .step-item:first-child {{ border-radius:8px 0 0 8px; }}
  .step-item:last-child  {{ border-radius:0 8px 8px 0; }}

  div[data-testid="stButton"] button[kind="primary"] {{
      background-color:#52B788 !important; border-color:#52B788 !important;
      color:white !important; font-weight:600 !important; border-radius:8px !important;
  }}
  div[data-testid="stButton"] button[kind="primary"]:hover {{
      background-color:#40916C !important; border-color:#40916C !important;
  }}
</style>
""", unsafe_allow_html=True)


# ── 세션 초기화 ───────────────────────────────
def init():
    defaults = {
        "step": 1,
        "api_key": "",
        "brand_name": "",
        "brand_color": "#4A90D9",
        "brand_category": "",
        "brand_usp": "",
        "brand_target": "",
        "brand_competitors": "",
        "brand_negative": "",
        "brand_focus": "",
        "research_summary": "",   # 웹 리서치 결과 요약
        "questions": [],
        "answers": {
            "off": {i: "" for i in range(1, 8)},
            "on":  {i: "" for i in range(1, 8)},
            "gem": {i: "" for i in range(1, 8)},
        },
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init()


# ── 헬퍼 ─────────────────────────────────────
def hex_to_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def get_client():
    return anthropic.Anthropic(api_key=st.session_state.api_key)

def check_mention(text: str, brand_name: str):
    """브랜드 언급 여부 + 맥락 반환"""
    if not text.strip():
        return None, ""
    keywords = [brand_name, brand_name.replace(" ", ""), brand_name[:2]]
    text_lower = text.lower()
    for kw in keywords:
        if kw.lower() in text_lower:
            idx = text_lower.find(kw.lower())
            start = max(0, idx - 80)
            end   = min(len(text), idx + 150)
            context = "…" + text[start:end].strip() + "…"
            return True, context
    return False, ""


# ── STEP 진행바 ───────────────────────────────
def render_stepbar():
    steps  = ["① 브랜드 입력", "② 질문 생성", "③ 질문 확인", "④ 답변 수집", "⑤ Excel 저장"]
    cur    = st.session_state.step
    html   = '<div class="step-bar">'
    for i, label in enumerate(steps, 1):
        cls = "step-active" if i == cur else ("step-done" if i < cur else "step-todo")
        html += f'<div class="step-item {cls}">{label}</div>'
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


# ── 앱 헤더 ──────────────────────────────────
bn = st.session_state.brand_name or "GEO Scan"
bc = st.session_state.brand_color
st.markdown(f"""
<div class="app-header">
  <div class="label">CREAMWORKS · GEO Scan v3.0</div>
  <div class="title">{bn} AI 진단 시스템</div>
  <div class="subtitle">브랜드 정보 입력 → 웹 리서치 → 맞춤 질문 생성 → 답변 수집 → Excel 저장</div>
</div>
""", unsafe_allow_html=True)

render_stepbar()
st.markdown("<hr class='divider'>", unsafe_allow_html=True)


# ══════════════════════════════════════════════
# STEP 1: 브랜드 정보 입력
# ══════════════════════════════════════════════
if st.session_state.step == 1:
    st.markdown('<div class="step-label">STEP 1</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-title">브랜드 정보 입력</div>', unsafe_allow_html=True)
    st.markdown('<div class="cw-box">💜 입력한 정보 + 웹 리서치를 결합해서 브랜드 맞춤 질문을 설계합니다.<br>많이 입력할수록 질문 퀄리티가 높아집니다.</div>', unsafe_allow_html=True)

    # API Key
    st.markdown("**🔑 Anthropic API Key**")
    api_key = st.text_input(
        "api", value=st.session_state.api_key,
        type="password", placeholder="sk-ant-api03-…",
        label_visibility="collapsed"
    )
    st.session_state.api_key = api_key
    st.caption("API 키는 세션 내에서만 유지되며 저장되지 않습니다.")

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**브랜드명** *")
        st.session_state.brand_name = st.text_input(
            "브랜드명", value=st.session_state.brand_name,
            placeholder="예: 교촌치킨", label_visibility="collapsed")

        st.markdown("**카테고리** *")
        st.session_state.brand_category = st.text_input(
            "카테고리", value=st.session_state.brand_category,
            placeholder="예: 치킨 프랜차이즈", label_visibility="collapsed")

        st.markdown("**브랜드 공식 컬러 (HEX)**")
        st.session_state.brand_color = st.color_picker(
            "컬러", value=st.session_state.brand_color,
            label_visibility="collapsed")
        st.markdown(f"""
        <div style="display:flex;align-items:center;gap:10px;margin:6px 0 2px">
          <div style="width:28px;height:28px;background:{st.session_state.brand_color};border-radius:5px;border:1px solid #ddd;"></div>
          <span style="font-size:.8rem;color:#666">{st.session_state.brand_color} — 보고서에 자동 적용</span>
        </div>
        <div style="font-size:.76rem;color:#aaa;margin-bottom:6px">
          구글에 "{st.session_state.brand_name or '브랜드명'} 브랜드 컬러 HEX" 검색 후 입력<br>
          예) 교촌치킨 #F9BA15 · 스타벅스 #00704A · 크림웍스 #7C5CBF
        </div>
        """, unsafe_allow_html=True)

        st.markdown("**경쟁 브랜드**")
        st.session_state.brand_competitors = st.text_input(
            "경쟁사", value=st.session_state.brand_competitors,
            placeholder="예: BBQ, BHC, 굽네치킨", label_visibility="collapsed")

        st.markdown("**주요 타겟 소비자**")
        st.session_state.brand_target = st.text_input(
            "타겟", value=st.session_state.brand_target,
            placeholder="예: 남녀노소, 배달앱 중심", label_visibility="collapsed")

    with col2:
        st.markdown("**핵심 USP (차별점)** *")
        st.session_state.brand_usp = st.text_area(
            "USP", value=st.session_state.brand_usp, height=110,
            placeholder="예: 간장치킨 원조, 35년 업력, 붓질 조리법, 국내산 신선 닭 100%",
            label_visibility="collapsed")

        st.markdown("**부정 이미지 / 약점**")
        st.session_state.brand_negative = st.text_area(
            "약점", value=st.session_state.brand_negative, height=80,
            placeholder="예: 가격 인상 논란, 양이 적다는 인식",
            label_visibility="collapsed")

        st.markdown("**이번 진단에서 특히 확인하고 싶은 것** *")
        st.session_state.brand_focus = st.text_area(
            "포인트", value=st.session_state.brand_focus, height=80,
            placeholder="예: 가성비 방어, 신메뉴 인지도, 배달앱 순위",
            label_visibility="collapsed")

    st.caption("* 필수 항목")
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    col_btn, _ = st.columns([2, 3])
    with col_btn:
        go = st.button("🔍 웹 리서치 + 질문 생성", type="primary", use_container_width=True)

    if go:
        # 유효성 검사
        missing = []
        if not st.session_state.api_key:     missing.append("Anthropic API Key")
        if not st.session_state.brand_name:  missing.append("브랜드명")
        if not st.session_state.brand_category: missing.append("카테고리")
        if not st.session_state.brand_usp:   missing.append("핵심 USP")
        if not st.session_state.brand_focus: missing.append("강조 포인트")
        if missing:
            st.error(f"필수 항목을 입력해주세요: {', '.join(missing)}")
        else:
            st.session_state.step = 2
            st.rerun()


# ══════════════════════════════════════════════
# STEP 2: 웹 리서치 + 질문 생성
# ══════════════════════════════════════════════
elif st.session_state.step == 2:
    st.markdown('<div class="step-label">STEP 2</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-title">웹 리서치 + 질문 자동 생성</div>', unsafe_allow_html=True)

    brand = st.session_state.brand_name

    with st.spinner(f"Claude가 {brand}를 분석하고 질문을 설계하는 중입니다..."):
        try:
            client = get_client()

            # ── 프롬프트 ──────────────────────────────
            prompt = f"""당신은 GEO(Generative Engine Optimization) 전문 컨설턴트입니다.
아래 브랜드 정보를 깊이 분석해서, AI가 소비자 질문에 답할 때 이 브랜드가 언급되는지 진단하기 위한 질문 7개를 설계하세요.

[브랜드 정보]
- 브랜드명: {st.session_state.brand_name}
- 카테고리: {st.session_state.brand_category}
- 핵심 USP: {st.session_state.brand_usp}
- 주요 타겟: {st.session_state.brand_target or '미입력'}
- 경쟁 브랜드: {st.session_state.brand_competitors or '미입력'}
- 부정 이미지/약점: {st.session_state.brand_negative or '미입력'}
- 강조 포인트: {st.session_state.brand_focus}

[질문 설계 원칙]
1. 브랜드명이 절대 포함되면 안 됨 — 브랜드를 모르는 소비자가 AI에게 자연스럽게 묻는 질문
2. 실제 소비자가 쓰는 구어체 (격식체 금지)
3. 단계 배분: DISCOVER 2개 / CONSIDER 3개 / DECIDE 2개
4. 브랜드 USP가 드러날 수 있는 질문 우선
5. 부정 이미지 방어 질문 반드시 1개 포함
6. 경쟁사와 비교 포지션을 확인할 수 있는 질문 포함

[근거 데이터 작성 규칙 - 매우 중요]
- 각 질문마다 실제 기관/매체의 데이터 3개를 넣을 것
- "출처기관" 같은 템플릿 텍스트 절대 금지. 반드시 실제 기관명 사용
- content에는 구체적 수치/통계/트렌드 포함
- year는 2023~2026 사이 실제 연도

[check_point 작성 규칙]
- 반드시 구체적으로: "{st.session_state.brand_name}이(가) [구체적 맥락]으로 등장하는지 + [{st.session_state.brand_competitors or '경쟁사'} 대비 포지션]"
- "확인 포인트:" 접두어 없이 내용만 작성

⚠️ 반드시 아래 JSON만 출력 (다른 텍스트 없이):
[
  {{
    "num": 1,
    "question": "질문 내용 (구어체, 브랜드명 없이)",
    "type": "유형명 (예: 시그니처메뉴탐색)",
    "stage": "DISCOVER",
    "check_point": "확인 포인트 내용",
    "is_priority": false,
    "data": [
      {{"source": "실제기관명", "content": "구체적 수치 포함 데이터", "year": "2024"}},
      {{"source": "실제기관명", "content": "구체적 수치 포함 데이터", "year": "2025"}},
      {{"source": "실제기관명", "content": "구체적 수치 포함 데이터", "year": "2024"}}
    ]
  }}
]"""

            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=8000,
                messages=[{"role": "user", "content": prompt}]
            )

            raw = response.content[0].text.strip()

            # JSON 추출
            json_match = re.search(r"\[.*\]", raw, re.DOTALL)
            if json_match:
                questions = json.loads(json_match.group())
            else:
                questions = json.loads(raw)

            st.session_state.questions = questions
            st.session_state.step = 3
            st.rerun()

        except json.JSONDecodeError:
            st.error("질문 생성 중 파싱 오류가 발생했습니다. 다시 시도해주세요.")
            if st.button("← 다시 시도"):
                st.session_state.step = 1
                st.rerun()
        except Exception as e:
            st.error(f"오류: {e}")
            if st.button("← 브랜드 정보 수정"):
                st.session_state.step = 1
                st.rerun()


# ══════════════════════════════════════════════
# STEP 3: 질문 확인 / 수정
# ══════════════════════════════════════════════
elif st.session_state.step == 3:
    st.markdown('<div class="step-label">STEP 3</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-title">질문 확인 및 수정</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="cw-box">💜 Claude가 <b>{st.session_state.brand_name}</b> 맞춤으로 설계한 질문입니다. 내용을 확인하고 필요하면 수정하세요.<br>⚠️ 브랜드명이 포함된 질문은 사용하지 않습니다.</div>', unsafe_allow_html=True)

    br, bg, bb = hex_to_rgb(st.session_state.brand_color)
    brand_hex  = st.session_state.brand_color

    updated_qs = []
    for i, q in enumerate(st.session_state.questions):
        n     = i + 1
        stage = q.get("stage", "")
        qtype = q.get("type", "")
        check = q.get("check_point", "")
        data  = q.get("data", [])
        is_priority = q.get("is_priority", False)

        stage_color = {"DISCOVER": "#2E86C1", "CONSIDER": "#1E8449", "DECIDE": "#884EA0"}.get(stage, "#555")

        st.markdown(f"""
        <div class="q-card">
          <span class="q-badge" style="background:{brand_hex}">Q{n}</span>
          <span class="q-badge" style="background:{stage_color}">{stage}</span>
          <span style="font-size:.78rem;color:#666;margin-left:4px">{qtype}</span>
          {"&nbsp;⭐" if is_priority else ""}
        </div>
        """, unsafe_allow_html=True)

        new_q = st.text_input(
            f"Q{n} 질문", value=q.get("question", ""),
            key=f"q_{i}", label_visibility="visible"
        )
        new_cp = st.text_input(
            "확인 포인트", value=check,
            key=f"cp_{i}", label_visibility="visible"
        )

        # 근거 데이터 표시
        if data:
            with st.expander(f"📊 선정 근거 데이터 ({len(data)}개)", expanded=False):
                dcols = st.columns(len(data))
                for j, d_item in enumerate(data[:3]):
                    with dcols[j]:
                        st.markdown(f"""
                        <div style="background:#f8f9fa;border-radius:8px;padding:10px;font-size:.77rem">
                          <div style="font-weight:700;color:#555;margin-bottom:4px">{d_item.get('source','')} ({d_item.get('year','')})</div>
                          <div style="color:#333;line-height:1.5">{d_item.get('content','')}</div>
                        </div>""", unsafe_allow_html=True)

        updated_qs.append({**q, "question": new_q, "check_point": new_cp})
        st.markdown("<hr style='border:none;border-top:1px solid #f0f0f0;margin:6px 0'>", unsafe_allow_html=True)

    col_back, col_regen, col_confirm = st.columns([1, 1, 2])
    with col_back:
        if st.button("← 정보 수정", use_container_width=True):
            st.session_state.step = 1
            st.rerun()
    with col_regen:
        if st.button("🔄 재생성", use_container_width=True):
            st.session_state.step = 2
            st.rerun()
    with col_confirm:
        if st.button("✅ 질문 확정 → 답변 수집 시작", type="primary", use_container_width=True):
            st.session_state.questions = updated_qs
            # 답변 초기화
            st.session_state.answers = {
                "off": {i: "" for i in range(1, 8)},
                "on":  {i: "" for i in range(1, 8)},
                "gem": {i: "" for i in range(1, 8)},
            }
            st.session_state.step = 4
            st.rerun()


# ══════════════════════════════════════════════
# STEP 4: AI 답변 수집
# ══════════════════════════════════════════════
elif st.session_state.step == 4:
    st.markdown('<div class="step-label">STEP 4</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-title">AI 답변 수집</div>', unsafe_allow_html=True)

    brand     = st.session_state.brand_name
    brand_hex = st.session_state.brand_color

    # 진행 현황
    off_cnt = sum(1 for v in st.session_state.answers["off"].values() if v.strip())
    on_cnt  = sum(1 for v in st.session_state.answers["on"].values() if v.strip())
    gem_cnt = sum(1 for v in st.session_state.answers["gem"].values() if v.strip())
    total   = off_cnt + on_cnt + gem_cnt

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("GPT 검색OFF", f"{off_cnt}/7")
    c2.metric("GPT 검색ON",  f"{on_cnt}/7")
    c3.metric("Gemini",      f"{gem_cnt}/7")
    c4.metric("전체",        f"{total}/21")
    st.progress(total / 21)

    # 세팅 안내
    st.markdown("""
    <div class="warn-box">
      <b>⚠️ 실행 전 필수 세팅</b><br>
      · ChatGPT 검색OFF: 새 채팅 → 메모리 OFF + 검색 OFF<br>
      · ChatGPT 검색ON:  새 채팅 → 메모리 OFF + 검색 ON<br>
      · Gemini: 시크릿 모드 → gemini.google.com 접속 (로그아웃 상태)<br>
      · 각 AI에서 Q1→Q7 <b>동일한 채팅 내</b>에서 순서대로 연속 입력하세요.
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    # ── AI별 탭 ──────────────────────────────
    tabs = st.tabs(["🔵 ChatGPT 검색OFF", "🟢 ChatGPT 검색ON", "🟠 Gemini"])
    ai_keys   = ["off", "on", "gem"]
    ai_labels = ["ChatGPT 검색OFF", "ChatGPT 검색ON", "Gemini"]
    ai_instrs = [
        "새 채팅 시작 → 메모리 OFF, 검색 OFF 확인 → Q1~Q7 순서대로 입력 → 각 답변 전체 복사 후 붙여넣기",
        "새 채팅 시작 → 메모리 OFF, 검색 ON 확인 → Q1~Q7 순서대로 입력 → 각 답변 전체 복사 후 붙여넣기",
        "시크릿 모드 → gemini.google.com (로그아웃) → 새 채팅 → Q1~Q7 순서대로 입력 → 각 답변 전체 복사 후 붙여넣기",
    ]

    for tab, key, label, instr in zip(tabs, ai_keys, ai_labels, ai_instrs):
        with tab:
            done_cnt = sum(1 for v in st.session_state.answers[key].values() if v.strip())
            st.progress(done_cnt / 7, text=f"{done_cnt}/7 완료")
            st.markdown(f'<div class="info-box">📌 {instr}</div>', unsafe_allow_html=True)

            for i, q in enumerate(st.session_state.questions):
                n     = i + 1
                q_txt = q.get("question", "")
                is_priority = q.get("is_priority", False)

                # 현재 저장 상태
                cur_val = st.session_state.answers[key].get(n, "")
                saved   = bool(cur_val.strip())

                c_q, c_badge = st.columns([6, 1])
                with c_q:
                    st.markdown(f"""
                    <div style="margin:10px 0 4px;font-weight:600;font-size:.88rem">
                      <span style="background:{brand_hex};color:white;padding:1px 8px;border-radius:5px;font-size:.76rem">Q{n}</span>
                      {"&nbsp;⭐" if is_priority else ""}
                      &nbsp; {q_txt}
                    </div>
                    """, unsafe_allow_html=True)
                    # 복사용 코드블록
                    st.code(q_txt, language=None)
                with c_badge:
                    st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
                    if saved:
                        st.markdown('<span class="mention-yes">✓ 저장됨</span>', unsafe_allow_html=True)
                    else:
                        st.markdown('<span class="mention-empty">미입력</span>', unsafe_allow_html=True)

                def make_cb(k, num):
                    def _cb():
                        val = st.session_state.get(f"ta_{k}_{num}", "")
                        st.session_state.answers[k][num] = val
                    return _cb

                st.text_area(
                    f"답변",
                    value=st.session_state.answers[key][n],
                    key=f"ta_{key}_{n}",
                    height=130,
                    placeholder=f"{label} 답변을 전체 복사해서 붙여넣으세요. 요약 없이 원문 전체를 넣어주세요.",
                    label_visibility="collapsed",
                    on_change=make_cb(key, n)
                )
                st.markdown("<hr style='border:none;border-top:1px solid #f4f4f4;margin:2px 0'>", unsafe_allow_html=True)

    # ── B2A 실시간 현황 ───────────────────────
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)
    st.markdown("#### 📊 B2A 현황 (실시간)")

    hcols = st.columns([4, 1, 1, 1])
    hcols[0].caption("질문")
    hcols[1].caption("GPT OFF")
    hcols[2].caption("GPT ON")
    hcols[3].caption("Gemini")

    for i, q in enumerate(st.session_state.questions):
        n = i + 1
        row = st.columns([4, 1, 1, 1])
        q_txt = q.get("question", "")
        row[0].write(f"Q{n}. {q_txt[:38]}{'…' if len(q_txt)>38 else ''}")
        for col, k in zip(row[1:], ["off", "on", "gem"]):
            ans = st.session_state.answers[k].get(n, "")
            mentioned, _ = check_mention(ans, brand)
            with col:
                if not ans.strip():
                    st.markdown('<span class="mention-empty">—</span>', unsafe_allow_html=True)
                elif mentioned:
                    st.markdown('<span class="mention-yes">✅</span>', unsafe_allow_html=True)
                else:
                    st.markdown('<span class="mention-no">❌</span>', unsafe_allow_html=True)

    # ── 하단 버튼 ─────────────────────────────
    st.markdown("<hr class='divider'>", unsafe_allow_html=True)
    col_back, col_save = st.columns([1, 3])
    with col_back:
        if st.button("← 질문 수정", use_container_width=True):
            st.session_state.step = 3
            st.rerun()
    with col_save:
        if st.button("📥 Excel 저장하기", type="primary", use_container_width=True):
            if total == 0:
                st.error("최소 1개 이상의 답변을 입력해주세요.")
            else:
                st.session_state.step = 5
                st.rerun()


# ══════════════════════════════════════════════
# STEP 5: Excel 생성 + 저장
# ══════════════════════════════════════════════
elif st.session_state.step == 5:
    st.markdown('<div class="step-label">STEP 5</div>', unsafe_allow_html=True)
    st.markdown('<div class="step-title">Excel 저장</div>', unsafe_allow_html=True)

    brand     = st.session_state.brand_name
    brand_hex = st.session_state.brand_color.lstrip("#")
    br, bg_c, bb = hex_to_rgb(st.session_state.brand_color)

    ai_keys   = ["off", "on", "gem"]
    ai_labels = ["ChatGPT 검색OFF", "ChatGPT 검색ON", "Gemini"]

    # ── 최종 B2A 요약 ─────────────────────────
    st.markdown("#### 📊 최종 B2A 결과")
    total_yes = 0
    total_q   = len(st.session_state.questions)

    for i, q in enumerate(st.session_state.questions):
        n   = i + 1
        row = st.columns([4, 1, 1, 1])
        q_txt = q.get("question", "")
        row[0].markdown(f"**Q{n}.** {q_txt}" + (" ⭐" if q.get("is_priority") else ""))
        for col, k in zip(row[1:], ai_keys):
            ans = st.session_state.answers[k].get(n, "")
            mentioned, _ = check_mention(ans, brand)
            with col:
                if not ans.strip():
                    st.markdown('<span class="mention-empty">미입력</span>', unsafe_allow_html=True)
                elif mentioned:
                    total_yes += 1
                    st.markdown('<span class="mention-yes">✅ 언급됨</span>', unsafe_allow_html=True)
                else:
                    st.markdown('<span class="mention-no">❌ 미언급</span>', unsafe_allow_html=True)

    total_possible = total_q * 3
    filled = sum(1 for k in ai_keys for n in range(1,8) if st.session_state.answers[k].get(n,"").strip())
    score_pct = round(total_yes / max(filled, 1) * 100)

    c1, c2, c3 = st.columns(3)
    c1.metric("B2A 언급 수", f"{total_yes} / {filled}건")
    c2.metric("언급률", f"{score_pct}%")
    c3.metric("수집 완료", f"{filled}/21건")

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    # ── Excel 생성 ────────────────────────────
    def make_excel() -> bytes:
        wb = openpyxl.Workbook()
        now_str = datetime.now().strftime("%Y.%m.%d")

        # 공통 스타일 헬퍼
        BRAND_HEX  = f"FF{brand_hex.upper()}"
        CW_HEX     = "FF7030A0"
        CW_MID     = "FF7C5CBF"
        CW_LIGHT   = "FFEADCF4"
        WHITE_HEX  = "FFFFFFFF"
        ODD_HEX    = "FFF5F2ED"
        EVEN_HEX   = "FFFFFFFF"
        GREEN_HEX  = "FFEAFAF1"
        RED_HEX    = "FFFDECEA"

        thin = Side(style="thin", color="DDDDDD")
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hf(color):  return PatternFill("solid", fgColor=color)
        def fnt(bold=False, color="FF1A1A1A", size=10):
            return Font(name="맑은 고딕", size=size, bold=bold, color=color)
        def aln(h="left", v="center", wrap=True):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def cell(ws, row, col, value, fill=None, bold=False, color="FF1A1A1A",
                 size=10, h="left", wrap=True, row_h=None):
            c = ws.cell(row=row, column=col, value=value)
            if fill: c.fill = fill
            c.font      = fnt(bold, color, size)
            c.alignment = aln(h, "center", wrap)
            c.border    = bdr
            if row_h:
                ws.row_dimensions[row].height = row_h
            return c

        # ══ 시트1: 브랜드 정보 ══
        ws1 = wb.active
        ws1.title = "브랜드 정보"
        ws1.column_dimensions["A"].width = 20
        ws1.column_dimensions["B"].width = 65

        ws1.merge_cells("A1:B1")
        ws1["A1"].value = f"GEO Scan — {brand} 브랜드 정보  |  {now_str}"
        ws1["A1"].fill  = hf(CW_HEX)
        ws1["A1"].font  = fnt(True, "FFFFFFFF", 13)
        ws1["A1"].alignment = aln("left", "center")
        ws1.row_dimensions[1].height = 30

        fields = [
            ("브랜드명",   st.session_state.brand_name),
            ("카테고리",   st.session_state.brand_category),
            ("핵심 USP",  st.session_state.brand_usp),
            ("주요 타겟",  st.session_state.brand_target),
            ("경쟁 브랜드", st.session_state.brand_competitors),
            ("부정 이미지", st.session_state.brand_negative),
            ("강조 포인트", st.session_state.brand_focus),
            ("브랜드 컬러", st.session_state.brand_color),
            ("진단일",     now_str),
        ]
        for i, (lbl, val) in enumerate(fields):
            row = i + 2
            cell(ws1, row, 1, lbl, hf(CW_MID), True, "FFFFFFFF", 10, "center", row_h=32)
            cell(ws1, row, 2, val, hf(ODD_HEX if i%2==0 else EVEN_HEX), row_h=32)

        # ══ 시트2: AI 답변 수집 (핵심) ══
        ws2 = wb.create_sheet("AI 답변 수집")

        # 컬럼 너비
        widths = {"A":7,"B":32,"C":18,"D":10,"E":38,
                  "F":55,"G":55,"H":55,
                  "I":14,"J":14,"K":14,
                  "L":45,"M":45,"N":45}
        for col, w in widths.items():
            ws2.column_dimensions[col].width = w

        # 타이틀
        ws2.merge_cells("A1:N1")
        ws2["A1"].value     = f"AI 답변 수집 — {brand}  |  수집일: {now_str}"
        ws2["A1"].fill      = hf(CW_HEX)
        ws2["A1"].font      = fnt(True, "FFFFFFFF", 13)
        ws2["A1"].alignment = aln("left", "center")
        ws2.row_dimensions[1].height = 30

        # 그룹 헤더
        groups = [
            ("A2:E2", "질문 정보",          CW_MID),
            ("F2:H2", "AI 답변 원문 (전체)", "FF2E86C1"),
            ("I2:K2", "B2A 언급 여부",       "FF1E8449"),
            ("L2:N2", "언급 맥락 (자동 추출)","FF7D3C98"),
        ]
        for rng, label, color in groups:
            ws2.merge_cells(rng)
            start = rng.split(":")[0]
            ws2[start].value     = label
            ws2[start].fill      = hf(color)
            ws2[start].font      = fnt(True, "FFFFFFFF", 10)
            ws2[start].alignment = aln("center", "center")
            ws2.row_dimensions[2].height = 20

        # 컬럼 헤더
        hdrs = [
            "번호","질문","유형","단계","확인 포인트",
            "ChatGPT 검색OFF 답변","ChatGPT 검색ON 답변","Gemini 답변",
            "GPT OFF 언급","GPT ON 언급","Gemini 언급",
            "GPT OFF 맥락","GPT ON 맥락","Gemini 맥락",
        ]
        hdr_colors = [CW_MID]*5 + ["FF2E86C1"]*3 + ["FF1E8449"]*3 + ["FF7D3C98"]*3
        for ci, (h, hc) in enumerate(zip(hdrs, hdr_colors), 1):
            c = ws2.cell(row=3, column=ci, value=h)
            c.fill      = hf(hc)
            c.font      = fnt(True, "FFFFFFFF", 10)
            c.alignment = aln("center", "center", True)
            c.border    = bdr
        ws2.row_dimensions[3].height = 26

        # 데이터 행
        for qi, q in enumerate(st.session_state.questions):
            n      = qi + 1
            row    = qi + 4
            row_bg = ODD_HEX if qi%2==0 else EVEN_HEX
            is_pri = q.get("is_priority", False)
            if is_pri: row_bg = "FFFFFDE7"

            q_label = f"Q{n}" + (" ⭐" if is_pri else "")
            ans_off = st.session_state.answers["off"].get(n, "")
            ans_on  = st.session_state.answers["on"].get(n, "")
            ans_gem = st.session_state.answers["gem"].get(n, "")

            men_off, ctx_off = check_mention(ans_off, brand)
            men_on,  ctx_on  = check_mention(ans_on,  brand)
            men_gem, ctx_gem = check_mention(ans_gem, brand)

            base_data = [
                q_label,
                q.get("question",""),
                q.get("type",""),
                q.get("stage",""),
                q.get("check_point",""),
                ans_off,
                ans_on,
                ans_gem,
            ]
            for ci, val in enumerate(base_data, 1):
                c = ws2.cell(row=row, column=ci, value=val)
                c.fill      = hf(row_bg)
                c.border    = bdr
                if ci in [6,7,8]:  # 답변 원문
                    c.font      = fnt(size=9)
                    c.alignment = aln("left", "top", True)
                elif ci == 1:
                    c.font      = fnt(True, size=10)
                    c.alignment = aln("center", "center")
                else:
                    c.font      = fnt(size=10)
                    c.alignment = aln("center", "center", True)

            # 언급 여부 (I, J, K)
            for ci, (mentioned, ans) in enumerate([(men_off, ans_off), (men_on, ans_on), (men_gem, ans_gem)], 9):
                c = ws2.cell(row=row, column=ci)
                c.border = bdr
                if not ans.strip():
                    c.value     = "—"
                    c.fill      = hf(row_bg)
                    c.font      = fnt(color="FF999999", size=10)
                elif mentioned:
                    c.value     = "✅ 언급됨"
                    c.fill      = hf(GREEN_HEX)
                    c.font      = fnt(True, "FF155724", 10)
                else:
                    c.value     = "❌ 미언급"
                    c.fill      = hf(RED_HEX)
                    c.font      = fnt(True, "FFB43216", 10)
                c.alignment = aln("center", "center")

            # 언급 맥락 (L, M, N)
            for ci, ctx in enumerate([ctx_off, ctx_on, ctx_gem], 12):
                c = ws2.cell(row=row, column=ci, value=ctx)
                c.fill      = hf(row_bg)
                c.font      = fnt(size=9, color="FF555555")
                c.alignment = aln("left", "top", True)
                c.border    = bdr

            ws2.row_dimensions[row].height = 90

        ws2.freeze_panes = "A4"

        # ══ 시트3: B2A 분석 결과 ══
        ws3 = wb.create_sheet("B2A 분석 결과")
        ws3.column_dimensions["A"].width = 45
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 18
        ws3.column_dimensions["D"].width = 18

        ws3.merge_cells("A1:D1")
        ws3["A1"].value     = f"B2A 매트릭스 — {brand}  |  수집일: {now_str}"
        ws3["A1"].fill      = hf(CW_HEX)
        ws3["A1"].font      = fnt(True, "FFFFFFFF", 13)
        ws3["A1"].alignment = aln("left", "center")
        ws3.row_dimensions[1].height = 30

        for ci, h in enumerate(["질문", "GPT 검색OFF", "GPT 검색ON", "Gemini"], 1):
            c = ws3.cell(row=2, column=ci, value=h)
            c.fill      = hf(CW_MID)
            c.font      = fnt(True, "FFFFFFFF", 11)
            c.alignment = aln("center", "center")
            c.border    = bdr
        ws3.row_dimensions[2].height = 24

        for qi, q in enumerate(st.session_state.questions):
            n   = qi + 1
            row = qi + 3
            row_bg = ODD_HEX if qi%2==0 else EVEN_HEX
            is_pri = q.get("is_priority", False)

            q_label = f"Q{n}. {q.get('question','')}" + (" ⭐" if is_pri else "")
            c = ws3.cell(row=row, column=1, value=q_label)
            c.fill      = hf(row_bg)
            c.font      = fnt(size=10)
            c.alignment = aln("left", "center", True)
            c.border    = bdr

            for ci, k in enumerate(ai_keys, 2):
                ans       = st.session_state.answers[k].get(n, "")
                mentioned, _ = check_mention(ans, brand)
                cell_obj  = ws3.cell(row=row, column=ci)
                cell_obj.border    = bdr
                cell_obj.alignment = aln("center", "center")
                if not ans.strip():
                    cell_obj.value = "—"
                    cell_obj.fill  = hf(row_bg)
                    cell_obj.font  = fnt(color="FF999999", size=11)
                elif mentioned:
                    cell_obj.value = "✅ 언급됨"
                    cell_obj.fill  = hf(GREEN_HEX)
                    cell_obj.font  = fnt(True, "FF155724", 11)
                else:
                    cell_obj.value = "❌ 미언급"
                    cell_obj.fill  = hf(RED_HEX)
                    cell_obj.font  = fnt(True, "FFB43216", 11)

            ws3.row_dimensions[row].height = 34

        # 총점 행
        sr = len(st.session_state.questions) + 4
        ws3.merge_cells(f"A{sr}:D{sr}")
        ws3[f"A{sr}"].value     = f"B2A 총점: {total_yes} / {filled}건  ({score_pct}%  |  수집완료: {filled}/21)"
        ws3[f"A{sr}"].fill      = hf(CW_LIGHT)
        ws3[f"A{sr}"].font      = Font(name="맑은 고딕", size=12, bold=True, color=CW_HEX)
        ws3[f"A{sr}"].alignment = aln("center", "center")
        ws3.row_dimensions[sr].height = 28

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ── 다운로드 버튼 ─────────────────────────
    st.markdown('<div class="cw-box">💜 아래 버튼으로 Excel을 다운로드하고, <b>클로드 프로젝트에 업로드</b>하세요.<br>클로드가 AI 답변 원문 전체를 분석해서 깊이 있는 제안서를 작성합니다.</div>', unsafe_allow_html=True)

    excel_data = make_excel()
    now_tag    = datetime.now().strftime("%Y%m%d_%H%M")
    filename   = f"{brand}_GEO_전체데이터_{now_tag}.xlsx"

    st.download_button(
        label="📥 Excel 다운로드",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )

    st.markdown("<hr class='divider'>", unsafe_allow_html=True)

    st.markdown("#### 📋 클로드 프로젝트 업로드 후 할 일")
    st.markdown(f"""
    1. **Excel 파일**을 클로드 프로젝트에 업로드
    2. 클로드에게: **"Excel 업로드했어. STEP 2 심층분석 진행해줘"**
    3. 클로드가 답변 원문 전체를 분석 → 제안서 작성

    > 💡 Excel의 **'AI 답변 수집'** 시트에 GPT/Gemini 답변 원문이 모두 저장되어 있어,
    > 클로드가 직접 인용하면서 깊이 있는 분석을 할 수 있습니다.
    """)

    col_back, col_new = st.columns([1, 3])
    with col_back:
        if st.button("← 답변 추가 수집", use_container_width=True):
            st.session_state.step = 4
            st.rerun()
    with col_new:
        if st.button("🔄 새 브랜드 진단 시작", use_container_width=True):
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()
